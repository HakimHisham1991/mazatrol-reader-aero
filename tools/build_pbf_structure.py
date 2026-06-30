"""Build pbf_structure.xlsx and pbf_structure.xml from SAMPLE_NC_PROGRAM/PBF HTML exports."""

from __future__ import annotations

import re
import shutil
import xml.etree.ElementTree as ET
from collections import defaultdict
from html import unescape
from html.parser import HTMLParser
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
PBF_DIR = ROOT / "SAMPLE_NC_PROGRAM" / "PBF"
QTS_XML = ROOT / "qts200m.xml"
OUT_XLSX = ROOT / "pbf_structure.xlsx"
OUT_XML = ROOT / "pbf_structure.xml"

# Provisional binary unit-type IDs (validate against real .PBF when available).
UNIT_IDS: dict[str, int] = {
    "MAT": 1,
    "WPC-": 2,
    "OFFSET": 3,
    "END": 4,
    "SUB PROGRAM": 5,
    "MANL PRG": 6,
    "M-CODE": 7,
    "HEAD": 19,
    "BAR": 48,
    "FACING": 51,
    "THREAD": 52,
    "T.GROOVE": 53,
    "T.DRILL": 54,
    "T.TAP": 55,
    "POCKET": 99,
    "DRILLING": 57,
    "LINE": 58,
    "TRANSFER": 59,
}

SUB_UNIT_IDS: dict[str, int] = {
    "SNo-MANL": 161,
    "SNo-BAR": 180,
    "SNo-POCKET": 178,
    "SNo-DRILLING": 182,
    "FIG-BAR": 168,
    "FIG-FACING": 172,
    "FIG-POCKET": 201,
    "FIG-DRILLING": 174,
}

PARAM_TYPES: dict[str, str] = {
    "MAT.": "text",
    "Material": "text",
    "OD-MAX": "readData",
    "OD": "readData",
    "ID-MIN": "readData",
    "ID": "readData",
    "LENGTH": "readData",
    "Length": "readData",
    "WORK FACE": "readData",
    "Workface": "readData",
    "ATC MODE": "readFullNumber2B",
    "RPM": "wholeNumber",
    "PART": "partType",
    "PART ": "partType",
    "CPT-X": "readData",
    "CPT-Z": "readData",
    "FIN-X": "readData",
    "FIN-Z": "readData",
    "MODE": "readPattern",
    "POS-C": "readData",
    "SRV-A": "readData",
    "BTM": "readData",
    "WAL": "readData",
    "FIN-A": "readData",
    "FIN-R": "readData",
    "INTER-R": "readData",
    "CHMF": "readData",
    "DIA": "readData",
    "DEPTH": "readData",
    "PAT.": "readPattern",
    "HEAD": "readFullNumber2B",
    "SPDL": "readPattern",
    "CONTI.": "readFullNumber2B",
    "REPEAT": "readFullNumber2B",
    "SHIFT": "readFullNumber2B",
    "NUMBER": "readFullNumber2B",
    "ATC": "readFullNumber2B",
    "RETURN": "readFullNumber2B",
    "WORK No.": "NA",
    "EXECTE": "readFullNumber2B",
    "TOOL": "text",
    "NOM-Ø": "readData",
    "NOM.": "readData",
    "No.": "readFullNumber2B",
    "No": "readFullNumber2B",
    "CHANGE-PT": "readPattern",
    "PTN": "readPattern",
    "SHIFT-Z": "readData",
    "R/x": "readData",
    "C/y": "readData",
    "SPT-X": "readData",
    "SPT-Z": "readData",
    "SPT-R/x": "readData",
    "SPT-C/y": "readData",
    "FPT-X": "readData",
    "FPT-Z": "readData",
    "ATTRIB": "readPattern",
    "APRCH-1": "readData",
    "APRCH-2": "readData",
    "PK-DEP": "readData",
    "DEP-A": "readData",
    "WID-R": "readData",
    "C-SP": "readFullNumber2B",
    "FR": "readData",
    "M": "readFullNumber2B",
    "G1": "readFullNumber2B",
    "G2": "readFullNumber2B",
    "DATA 1": "readLetter",
    "DATA 2": "readLetter",
    "DATA 3": "readLetter",
    "DATA 4": "readLetter",
    "DATA 5": "readLetter",
    "DATA 6": "readLetter",
    "S": "readPattern",
    "M/B": "readFullNumber2B",
    "M1": "readFullNumber2B",
    "M2": "readFullNumber2B",
    "M3": "readFullNumber2B",
    "M4": "readFullNumber2B",
    "M5": "readFullNumber2B",
    "M6": "readFullNumber2B",
    "M7": "readFullNumber2B",
    "M8": "readFullNumber2B",
    "M9": "readFullNumber2B",
    "M10": "readFullNumber2B",
    "M11": "readFullNumber2B",
    "M12": "readFullNumber2B",
}


class TableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.tables: list[list[list[str]]] = []
        self._in_table = False
        self._in_row = False
        self._in_cell = False
        self._current_table: list[list[str]] = []
        self._current_row: list[str] = []
        self._cell_parts: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag == "table":
            self._in_table = True
            self._current_table = []
        elif tag == "tr" and self._in_table:
            self._in_row = True
            self._current_row = []
        elif tag in ("td", "th") and self._in_row:
            self._in_cell = True
            self._cell_parts = []

    def handle_endtag(self, tag: str) -> None:
        if tag in ("td", "th") and self._in_cell:
            self._in_cell = False
            self._current_row.append(unescape("".join(self._cell_parts)).strip())
        elif tag == "tr" and self._in_row:
            self._in_row = False
            if self._current_row:
                self._current_table.append(self._current_row)
        elif tag == "table" and self._in_table:
            self._in_table = False
            if self._current_table:
                self.tables.append(self._current_table)

    def handle_data(self, data: str) -> None:
        if self._in_cell:
            self._cell_parts.append(data)


def strip_html(text: str) -> str:
    return unescape(re.sub(r"<[^>]+>", "", text))


def infer_unit_from_header(header_cols: list[str], row_parts: list[str]) -> tuple[str, list[str], list[str]]:
    """Derive unit name and column layout from a Mazatrol PRE header row."""
    header_upper = [c.upper().replace(".", "") for c in header_cols]

    if header_upper[0] != "UNO":
        return "?", header_cols[1:], row_parts

    if len(header_upper) > 1 and header_upper[1] == "MAT":
        return "MAT", header_cols[2:], row_parts[1:]

    if "TOOL" in header_upper and "NOM" in "".join(header_upper):
        return "MANL PRG", header_cols[2:], row_parts[2:]

    if "M1" in header_upper and "M2" in header_upper:
        return "M-CODE", header_cols[2:], row_parts[2:]

    if "CONTI" in header_upper and "REPEAT" in header_upper:
        return "END", header_cols[2:], row_parts[2:]

    if "PAT" in header_upper and "HEAD" in header_upper:
        return "HEAD", header_cols[2:], row_parts[2:]

    if "MODE" in header_upper and "DIA" in header_upper and "DEPTH" in header_upper:
        return "DRILLING", header_cols[2:], row_parts[2:]

    if "MODE" in header_upper and "POS-C" in header_upper:
        return "POCKET", header_cols[2:], row_parts[2:]

    # UNo. UNIT PART ... (BAR, FACING, etc.)
    if len(header_upper) > 2 and header_upper[1] == "UNIT" and header_upper[2] == "PART":
        unit_name = row_parts[1] if len(row_parts) > 1 else "?"
        return unit_name, header_cols[3:], row_parts[2:]

    if len(header_upper) > 1 and header_upper[1] == "UNIT":
        unit_name = row_parts[1] if len(row_parts) > 1 else "?"
        if unit_name == "MANL" and len(row_parts) > 2 and row_parts[2] == "PRG":
            return "MANL PRG", header_cols[2:], row_parts[3:]
        return unit_name, header_cols[2:], row_parts[2:]

    return "?", header_cols[1:], row_parts[1:]


def parse_pre_tables(path: Path) -> list[dict]:
    """Parse legacy FONT/PRE Mazatrol HTML exports."""
    raw = path.read_text(encoding="utf-8", errors="replace")
    if "UNo." not in raw:
        raw = path.read_text(encoding="windows-1252", errors="replace")
    lines = [strip_html(line).rstrip() for line in raw.splitlines() if strip_html(line).strip()]

    blocks: list[dict] = []
    i = 0
    while i < len(lines):
        line = lines[i]
        if not line.strip().startswith("UNo."):
            i += 1
            continue

        header_cols = re.sub(r"\s+", " ", line.strip()).split()
        i += 1
        while i < len(lines) and not lines[i].strip():
            i += 1
        if i >= len(lines):
            break

        row = lines[i]
        i += 1
        row_parts = row.split()
        if not row_parts or not row_parts[0].strip().isdigit():
            continue

        unit_name, data_cols, values = infer_unit_from_header(header_cols, row_parts)
        if unit_name == "?" or unit_name.isdigit():
            continue

        block = {
            "file": path.name,
            "unit_name": unit_name,
            "header": header_cols,
            "data_cols": data_cols,
            "values": values,
            "sub_units": [],
        }

        while i < len(lines):
            sub = lines[i].strip()
            if sub.startswith("UNo."):
                break
            if sub.startswith("SNo.") or sub.startswith("FIG"):
                sub_header = re.sub(r"\s+", " ", sub).split()
                i += 1
                sub_rows: list[list[str]] = []
                while i < len(lines):
                    row_line = lines[i].strip()
                    if row_line.startswith("UNo.") or row_line.startswith("SNo.") or row_line.startswith("FIG"):
                        break
                    if row_line:
                        sub_rows.append(re.split(r"\s{2,}", row_line))
                    i += 1
                block["sub_units"].append({"kind": sub_header[0], "header": sub_header, "rows": sub_rows})
                continue
            i += 1

        blocks.append(block)
    return blocks


def parse_html_tables(path: Path) -> list[dict]:
    parser = TableParser()
    parser.feed(path.read_text(encoding="utf-8", errors="replace"))

    blocks: list[dict] = []
    pending: dict | None = None

    for table in parser.tables:
        if not table:
            continue
        header = [c.strip() for c in table[0]]
        if not header or header[0] not in {"UNo.", "FIG", "SNo.", ""}:
            if header and header[0] == "FIG" and pending:
                pending["sub_units"].append(
                    {"kind": "FIG", "header": header, "rows": table[1:]}
                )
            elif header and (header[0] in {"", "SNo."} or header[1] in {"SNo."}) and pending:
                pending["sub_units"].append(
                    {"kind": "SNo", "header": header, "rows": table[1:]}
                )
            continue

        if header[0] != "UNo.":
            continue

        for row in table[1:]:
            if not row or not row[0].strip().isdigit():
                continue
            if header[1].upper().replace(".", "") == "MAT":
                unit_name = "MAT"
                data_cols = header[2:]
                values = row[1:]
            else:
                unit_name = row[1].strip() if len(row) > 1 else "?"
                data_cols = header[2:]
                values = row[2:]

            pending = {
                "file": path.name,
                "unit_name": unit_name,
                "header": header,
                "data_cols": data_cols,
                "values": values,
                "sub_units": [],
            }
            blocks.append(pending)
    return blocks


def collect_blocks() -> list[dict]:
    blocks: list[dict] = []
    for path in sorted(PBF_DIR.glob("*.html")):
        text = path.read_text(encoding="utf-8", errors="replace")
        if "<table" in text.lower():
            blocks.extend(parse_html_tables(path))
        else:
            blocks.extend(parse_pre_tables(path))
    return blocks


def infer_param_type(name: str) -> str:
    clean = name.strip().replace("  ", " ")
    if clean in PARAM_TYPES:
        return PARAM_TYPES[clean]
    if clean.endswith("."):
        return PARAM_TYPES.get(clean[:-1] + ".", "readData")
    return "readData"


def load_qts_offsets() -> dict[tuple[str, str], int]:
    offsets: dict[tuple[str, str], int] = {}
    if not QTS_XML.is_file():
        return offsets
    root = ET.parse(QTS_XML).getroot()
    for unit in root.findall("unit"):
        uid = unit.get("name", "")
        for param in unit.findall("parameter"):
            offsets[(uid, param.get("name", ""))] = int(param.get("pos", "0"))
    return offsets


def build_records(blocks: list[dict]) -> tuple[list[dict], list[dict], list[dict]]:
    qts_offsets = load_qts_offsets()
    units: dict[str, dict] = {}
    params: list[dict] = []
    sub_params: list[dict] = []

    for block in blocks:
        unit_name = block["unit_name"]
        uid = UNIT_IDS.get(unit_name, 0)
        if unit_name not in units:
            units[unit_name] = {
                "unit_id": uid,
                "unit_name": unit_name,
                "sample_files": set(),
                "instance_count": 0,
            }
        units[unit_name]["sample_files"].add(block["file"])
        units[unit_name]["instance_count"] += 1

        for col in block["data_cols"]:
            col = col.strip()
            if not col:
                continue
            ptype = infer_param_type(col)
            pos = qts_offsets.get((unit_name, col), qts_offsets.get((unit_name, col.rstrip(".")), 0))
            params.append(
                {
                    "unit_id": uid,
                    "unit_name": unit_name,
                    "parameter": col,
                    "display_type": ptype,
                    "binary_offset": pos,
                    "source": "qts200m.xml" if pos else "html-inferred",
                    "sample_file": block["file"],
                }
            )

        for sub in block["sub_units"]:
            kind = sub["kind"]
            parent = unit_name
            if kind == "FIG":
                sid = {
                    "POCKET": SUB_UNIT_IDS["FIG-POCKET"],
                    "DRILLING": SUB_UNIT_IDS["FIG-DRILLING"],
                    "BAR": SUB_UNIT_IDS["FIG-BAR"],
                    "FACING": SUB_UNIT_IDS["FIG-FACING"],
                }.get(parent, SUB_UNIT_IDS["FIG-BAR"])
                sub_name = f"FIG ({parent})"
            else:
                sid = {
                    "POCKET": SUB_UNIT_IDS["SNo-POCKET"],
                    "DRILLING": SUB_UNIT_IDS["SNo-DRILLING"],
                    "BAR": SUB_UNIT_IDS["SNo-BAR"],
                    "MANL PRG": SUB_UNIT_IDS["SNo-MANL"],
                }.get(parent, SUB_UNIT_IDS["SNo-BAR"])
                sub_name = f"SNo ({parent})"

            header = [h for h in sub["header"] if h and h not in {"FIG", "SNo."}]
            for col in header:
                ptype = infer_param_type(col)
                sub_params.append(
                    {
                        "sub_unit_id": sid,
                        "sub_unit_name": sub_name,
                        "parent_unit": parent,
                        "parameter": col,
                        "display_type": ptype,
                        "binary_offset": 0,
                        "source": "html-inferred",
                        "sample_file": block["file"],
                    }
                )

    unit_rows = []
    for name in sorted(units, key=lambda n: units[n]["unit_id"] or 999):
        u = units[name]
        unit_rows.append(
            {
                "unit_id": u["unit_id"],
                "unit_name": name,
                "instance_count": u["instance_count"],
                "sample_files": ", ".join(sorted(u["sample_files"])),
                "notes": "Provisional ID — validate with binary .PBF" if u["unit_id"] == 0 else "",
            }
        )

    def dedupe(rows: list[dict], key_fields: list[str]) -> list[dict]:
        seen: set[tuple] = set()
        out: list[dict] = []
        for row in rows:
            key = tuple(row[k] for k in key_fields)
            if key in seen:
                continue
            seen.add(key)
            out.append(row)
        return out

    return (
        unit_rows,
        dedupe(params, ["unit_name", "parameter"]),
        dedupe(sub_params, ["sub_unit_name", "parameter"]),
    )


def write_xlsx(unit_rows: list[dict], param_rows: list[dict], sub_rows: list[dict]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Units"
    ws.append(["unit_id", "unit_name", "instance_count", "sample_files", "notes"])
    for row in unit_rows:
        ws.append([row["unit_id"], row["unit_name"], row["instance_count"], row["sample_files"], row["notes"]])

    ws2 = wb.create_sheet("Parameters")
    ws2.append(["unit_id", "unit_name", "parameter", "display_type", "binary_offset", "source"])
    for row in param_rows:
        ws2.append(
            [row["unit_id"], row["unit_name"], row["parameter"], row["display_type"], row["binary_offset"], row["source"]]
        )

    ws3 = wb.create_sheet("SubUnits")
    ws3.append(["sub_unit_id", "sub_unit_name", "parent_unit", "parameter", "display_type", "binary_offset", "source"])
    for row in sub_rows:
        ws3.append(
            [
                row["sub_unit_id"],
                row["sub_unit_name"],
                row["parent_unit"],
                row["parameter"],
                row["display_type"],
                row["binary_offset"],
                row["source"],
            ]
        )

    ws4 = wb.create_sheet("About")
    ws4["A1"] = "PBF structure reverse-engineered from SAMPLE_NC_PROGRAM/PBF HTML exports"
    ws4["A1"].font = Font(bold=True)
    ws4["A3"] = "Binary offsets copied from qts200m.xml where units overlap; others are inferred from display columns."
    ws4["A4"] = "Unit type IDs 13, 56, 57, 169, 181, 182, 174 are provisional until validated against binary .PBF files."

    wb.save(OUT_XLSX)


def append_unit(root: ET.Element, unit_id: int, name: str, params: list[tuple[str, int, str]]) -> None:
    unit = ET.SubElement(root, "unit", id=str(unit_id), name=name)
    for pname, pos, ptype in params:
        ET.SubElement(unit, "parameter", name=pname, pos=str(pos), type=ptype)


def write_xml(param_rows: list[dict], sub_rows: list[dict]) -> None:
    shutil.copy(QTS_XML, OUT_XML)
    root = ET.parse(OUT_XML).getroot()
    assert root is not None

    # Extend MANL PRG
    for unit in root.findall("unit"):
        if unit.get("id") == "6":
            existing = {p.get("name") for p in unit.findall("parameter")}
            for name, pos, ptype in [("TOOL", 84, "text"), ("NOM-Ø", 64, "readData"), ("No.", 68, "readFullNumber2B")]:
                if name not in existing:
                    ET.SubElement(unit, "parameter", name=name, pos=str(pos), type=ptype)

    append_unit(
        root,
        13,
        "HEAD",
        [("PAT.", 64, "readPattern"), ("HEAD", 68, "readFullNumber2B"), ("SPDL", 72, "readPattern")],
    )
    append_unit(
        root,
        56,
        "POCKET",
        [
            ("MODE", 64, "readPattern"),
            ("POS-C", 36, "readData"),
            ("SRV-A", 40, "readData"),
            ("BTM", 44, "readData"),
            ("WAL", 48, "readData"),
            ("FIN-A", 52, "readData"),
            ("FIN-R", 56, "readData"),
            ("INTER-R", 60, "readData"),
            ("CHMF", 76, "readData"),
        ],
    )
    append_unit(
        root,
        57,
        "DRILLING",
        [
            ("MODE", 64, "readPattern"),
            ("POS-C", 36, "readData"),
            ("DIA", 40, "readData"),
            ("DEPTH", 44, "readData"),
            ("CHMF", 76, "readData"),
        ],
    )

    append_unit(
        root,
        169,
        "FIG",
        [
            ("PTN", 8, "readPattern"),
            ("SHIFT-Z", 40, "readData"),
            ("R/x", 36, "readData"),
            ("C/y", 44, "readData"),
            ("R/th", 56, "readData"),
            ("ATTRIB", 52, "readPattern"),
            ("FR", 64, "readData"),
        ],
    )
    append_unit(
        root,
        174,
        "FIG",
        [
            ("PTN", 8, "readPattern"),
            ("SPT-R/x", 36, "readData"),
            ("SPT-C/y", 44, "readData"),
            ("SPT-Z", 40, "readData"),
            ("NUM.", 48, "readData"),
            ("ANG", 52, "readData"),
        ],
    )
    append_unit(
        root,
        181,
        "SNo",
        [
            ("TOOL", 84, "text"),
            ("NOM-Ø", 64, "readData"),
            ("APRCH-1", 48, "readData"),
            ("APRCH-2", 52, "readData"),
            ("PK-DEP", 56, "readData"),
            ("DEP-A", 60, "readData"),
            ("WID-R", 76, "readData"),
            ("C-SP", 88, "readFullNumber2B"),
            ("FR", 92, "readData"),
            ("M", 24, "readFullNumber2B"),
        ],
    )
    append_unit(
        root,
        182,
        "SNo",
        [
            ("TOOL", 84, "text"),
            ("NOM-Ø", 64, "readData"),
            ("HOLE-Ø", 48, "readData"),
            ("HOLE-DEP", 52, "readData"),
            ("PRE-DIA", 56, "readData"),
            ("PRE-DEP", 60, "readData"),
            ("DEPTH", 76, "readData"),
            ("C-SP", 88, "readFullNumber2B"),
            ("FR", 92, "readData"),
            ("M", 24, "readFullNumber2B"),
        ],
    )

    tree = ET.ElementTree(root)
    ET.indent(tree, space="    ")
    tree.write(OUT_XML, encoding="utf-8", xml_declaration=True)


def main() -> None:
    blocks = collect_blocks()
    unit_rows, param_rows, sub_rows = build_records(blocks)
    write_xlsx(unit_rows, param_rows, sub_rows)
    write_xml(param_rows, sub_rows)
    print(f"Parsed {len(blocks)} unit blocks from {len(list(PBF_DIR.glob('*.html')))} HTML files")
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_XML}")
    print(f"Distinct unit types: {len(unit_rows)}")


if __name__ == "__main__":
    main()
