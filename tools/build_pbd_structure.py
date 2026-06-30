"""Build pbd_structure.xlsx and pbd_structure.xml from SAMPLE_NC_PROGRAM/PBD HTML exports."""

from __future__ import annotations

import re
import shutil
import struct
import sys
import xml.etree.ElementTree as ET
from collections import Counter
from html import unescape
from html.parser import HTMLParser
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from tools.build_pbf_structure import (  # noqa: E402
    PARAM_TYPES,
    infer_param_type,
    strip_html,
)


def dedupe_rows(rows: list[dict], key_fields: list[str]) -> list[dict]:
    seen: set[tuple] = set()
    out: list[dict] = []
    for row in rows:
        key = tuple(row[k] for k in key_fields)
        if key in seen:
            continue
        seen.add(key)
        out.append(row)
    return out

PBD_DIR = ROOT / "SAMPLE_NC_PROGRAM" / "PBD"
QTS_XML = ROOT / "qts200m.xml"
OUT_XLSX = ROOT / "pbd_structure.xlsx"
OUT_XML = ROOT / "pbd_structure.xml"

START = 0xFC
UNIT_SIZE = 100

# Binary-validated unit type IDs (SAMPLE_NC_PROGRAM/PBD/*.PBD).
UNIT_IDS: dict[str, int] = {
    "MAT": 1,
    "WPC-": 2,
    "OFFSET": 3,
    "END": 4,
    "SUB PROGRAM": 5,
    "SUB PRO": 5,
    "MANL PRG": 6,
    "INDEX": 12,
    "DRILLING": 32,
    "REAMING": 35,
    "CIRC MIL": 38,
    "LINE CTR": 64,
    "LINE LFT": 66,
    "LINE IN": 68,
    "POCKET": 99,
}

SUB_UNIT_IDS: dict[str, int] = {
    "SNo-MANL": 161,
    "SNo-LINE": 177,
    "SNo-DRILLING": 176,
    "SNo-POCKET": 178,
    "FIG-LINE": 194,
    "FIG-DRILLING": 192,
    "FIG-POCKET": 193,
}

PBD_MULTI_MODE_ENUMS: list[tuple[str, int]] = [
    ("OFF", 1),
    ("OFFSET", 3),
]

PBD_PTN_ENUMS: dict[int, list[tuple[str, int]]] = {
    # FIG point geometry (drill/circ/ream) — HTML: PT
    192: [("PT", 1)],
    # FIG pocket geometry — HTML: SQR, etc.
    193: [("SQR", 16), ("LINE", 32), ("ARC", 33)],
    # FIG contour geometry — HTML: LINE, CW, CCW
    194: [("LINE", 32), ("CW", 33), ("CCW", 34)],
}

# PBD SNo TOOL: packed (byte+9 << 8 | byte+13) — validated against all PBD HTML samples.
PBD_TOOL_ENUMS: list[tuple[str, int]] = [
    ("DRILL", 512),
    ("FCE MILL", 3596),
    ("END MILL", 3840),
    ("END MILL", 3844),
    ("END MILL", 3852),
    ("BAL EMIL", 4876),
]

# Binary-validated parameter offsets for pbd_structure.xml generation.
BINARY_OFFSETS: dict[int, list[tuple[str, int, str]]] = {
    1: [
        ("MAT.", 84, "text"),
        ("INITIAL-Z", 40, "readData"),
        ("ATC MODE", 10, "readFullNumber2B"),
        ("MULTI MODE", 9, "readPattern"),
        ("MULTI FLAG", 9, "readPbdMultiFlag"),
        ("PITCH-X", 72, "readData"),
        ("PITCH-Y", 76, "readData"),
    ],
    160: [
        ("X", 36, "readData"),
        ("Y", 40, "readData"),
        ("th", 44, "readData"),
        ("Z", 48, "readData"),
    ],
    2: [
        ("ADD. WPC", 32, "readFullNumber2B"),
        ("X", 36, "readData"),
        ("Y", 40, "readData"),
        ("th", 44, "readData"),
        ("Z", 48, "readData"),
    ],
    12: [
        ("TURN POS X", 36, "readData"),
        ("TURN POS Y", 40, "readData"),
        ("TURN POS Z", 44, "readData"),
    ],
    32: [
        ("DIA", 36, "readData"),
        ("DEPTH", 40, "readData"),
        ("CHMF", 44, "readData"),
    ],
    35: [
        ("DIA", 36, "readData"),
        ("DEPTH", 40, "readData"),
        ("CHMF", 44, "readData"),
    ],
    38: [
        ("TORNAD", 32, "readFullNumber2B"),
        ("DIA", 36, "readData"),
        ("DEPTH", 40, "readData"),
        ("CHMF", 44, "readData"),
        ("BTM", 48, "readFullNumber2B"),
        ("PITCH1", 52, "readData"),
        ("PITCH2", 56, "readData"),
    ],
    64: [
        ("DEPTH", 36, "readData"),
        ("SRV-Z", 40, "readData"),
        ("SRV-R", 44, "readData"),
        ("RGH", 48, "readFullNumber2B"),
        ("FIN-Z", 52, "readData"),
    ],
    66: [
        ("DEPTH", 36, "readData"),
        ("SRV-Z", 40, "readData"),
        ("SRV-R", 44, "readData"),
        ("RGH", 48, "readFullNumber2B"),
        ("FIN-Z", 52, "readData"),
        ("FIN-R", 56, "readData"),
        ("INTER-R", 60, "readData"),
        ("CHMF", 64, "readData"),
    ],
    68: [
        ("DEPTH", 36, "readData"),
        ("SRV-Z", 40, "readData"),
        ("SRV-R", 44, "readData"),
        ("RGH", 48, "readFullNumber2B"),
        ("FIN-Z", 52, "readData"),
    ],
    99: [
        ("POS-C", 36, "readData"),
        ("SRV-A", 40, "readData"),
        ("BTM", 48, "readFullNumber2B"),
        ("WAL", 52, "readFullNumber2B"),
        ("FIN-A", 56, "readData"),
        ("FIN-R", 60, "readData"),
        ("INTER-R", 64, "readData"),
        ("CHMF", 68, "readData"),
    ],
    176: [
        ("TOOL", 9, "readPbdTool"),
        ("NOM-Ø", 36, "readData"),
        ("HOLE-Ø", 40, "readData"),
        ("HOLE-DEP", 44, "readData"),
        ("DEPTH", 52, "readData"),
        ("C-SP", 60, "readFullNumber2B"),
        ("FR", 64, "readData"),
        ("M", 24, "readFullNumber2B"),
        ("M", 26, "readFullNumber2B"),
        ("M", 32, "readFullNumber2B"),
    ],
    177: [
        ("TOOL", 9, "readPbdTool"),
        ("NOM-Ø", 36, "readData"),
        ("APRCH-X", 40, "readData"),
        ("APRCH-Y", 44, "readData"),
        ("DEP-Z", 48, "readData"),
        ("WID-R", 52, "readData"),
        ("C-SP", 60, "readFullNumber2B"),
        ("FR", 64, "readData"),
        ("M", 24, "readFullNumber2B"),
        ("M", 26, "readFullNumber2B"),
        ("M", 32, "readFullNumber2B"),
    ],
    178: [
        ("TOOL", 9, "readPbdTool"),
        ("NOM-Ø", 36, "readData"),
        ("PK-DEP", 48, "readData"),
        ("WID-R", 52, "readData"),
        ("C-SP", 60, "readFullNumber2B"),
        ("FR", 64, "readData"),
        ("M", 24, "readFullNumber2B"),
        ("M", 26, "readFullNumber2B"),
        ("M", 32, "readFullNumber2B"),
    ],
    192: [
        ("PTN", 8, "readPattern"),
        ("Z", 36, "readData"),
        ("X", 40, "readData"),
        ("Y", 44, "readData"),
    ],
    193: [
        ("PTN", 8, "readPattern"),
        ("R/x", 36, "readData"),
        ("C/y", 40, "readData"),
        ("R/th", 44, "readData"),
        ("SHIFT-Z", 48, "readData"),
        ("FR", 64, "readData"),
    ],
    194: [
        ("PTN", 8, "readPattern"),
        ("X", 40, "readData"),
        ("Y", 44, "readData"),
        ("R/th", 56, "readData"),
        ("I", 60, "readData"),
        ("J", 64, "readData"),
    ],
}

PBD_PARAM_TYPES = {
    **PARAM_TYPES,
    "INITIAL-Z": "readData",
    "MULTI MODE": "readPattern",
    "MULTI FLAG": "readPattern",
    "PITCH-X": "readData",
    "PITCH-Y": "readData",
    "TURN POS X": "readData",
    "TURN POS Y": "readData",
    "TURN POS Z": "readData",
    "ANGLEC": "readData",
    "ANGLEA": "readData",
    "ADD. WPC": "readFullNumber2B",
    "th": "readData",
    "SRV-Z": "readData",
    "SRV-R": "readData",
    "SRV-A": "readData",
    "FIN-Z": "readData",
    "FIN-R": "readData",
    "INTER-R": "readData",
    "START": "readPattern",
    "END": "readPattern",
    "TORNAD": "readFullNumber2B",
    "PITCH1": "readData",
    "PITCH2": "readData",
    "PRE-REAM": "readPattern",
    "CHP": "readFullNumber2B",
    "PRE-DIA": "readData",
    "BTM": "readFullNumber2B",
    "APRCH-X": "readData",
    "APRCH-Y": "readData",
    "TYPE": "readPattern",
    "ZFD": "readFullNumber2B",
    "DEP-Z": "readData",
    "WID-R": "readData",
    "RGH": "readFullNumber2B",
    "AN1": "readData",
    "AN2": "readData",
    "T1": "readData",
    "T2": "readData",
    "R/th": "readData",
    "CNR": "readData",
    "R-FEED": "readData",
    "PK-DEP": "readData",
}


def infer_pbd_unit_from_header(header_cols: list[str], row_parts: list[str]) -> tuple[str, list[str], list[str]]:
    header_upper = [c.upper().replace(".", "") for c in header_cols]

    if header_upper[0] != "UNO":
        return "?", header_cols[1:], row_parts

    if len(header_upper) > 1 and header_upper[1] in {"MAT", "MATERIAL"}:
        return "MAT", header_cols[2:], row_parts[1:]

    if "INITIAL-Z" in header_upper or (len(header_upper) > 2 and header_upper[2] == "INITIAL-Z"):
        return "MAT", header_cols[2:], row_parts[1:]

    if "TURN POS X" in header_upper:
        unit_name = row_parts[1] if len(row_parts) > 1 else "INDEX"
        return unit_name.strip(), header_cols[2:], row_parts[2:]

    if "ADD. WPC" in header_upper or "ADD WPC" in " ".join(header_upper):
        return "WPC-", header_cols[2:], row_parts[2:]

    if "DIA" in header_upper and "DEPTH" in header_upper and "CHMF" in header_upper:
        if "PRE-REAM" in header_upper:
            unit_name = row_parts[1] if len(row_parts) > 1 else "REAMING"
            return unit_name.strip(), header_cols[2:], row_parts[2:]
        if "TORNAD" in header_upper:
            unit_name = row_parts[1] if len(row_parts) > 1 else "CIRC MIL"
            return unit_name.strip(), header_cols[2:], row_parts[2:]
        if "UNIT" in header_upper:
            unit_name = row_parts[1] if len(row_parts) > 1 else "DRILLING"
            return unit_name.strip(), header_cols[2:], row_parts[2:]

    if "SRV-Z" in header_upper and "SRV-R" in header_upper:
        unit_name = row_parts[1] if len(row_parts) > 1 else "LINE LFT"
        if len(row_parts) > 2 and row_parts[2] == "LFT":
            unit_name = "LINE LFT"
            return unit_name, header_cols[2:], row_parts[3:]
        if len(row_parts) > 2 and row_parts[2] == "CTR":
            return "LINE CTR", header_cols[2:], row_parts[3:]
        if len(row_parts) > 2 and row_parts[2] == "IN":
            return "LINE IN", header_cols[2:], row_parts[3:]
        return unit_name.strip(), header_cols[2:], row_parts[2:]

    if "POS-C" in header_upper or ("POS" in header_upper and "SRV-A" in header_upper):
        unit_name = row_parts[1] if len(row_parts) > 1 else "POCKET"
        return unit_name.strip(), header_cols[2:], row_parts[2:]

    if "TOOL" in header_upper and "NOM" in "".join(header_upper):
        return "MANL PRG", header_cols[2:], row_parts[2:]

    if "CONTI" in header_upper and "REPEAT" in header_upper:
        return "END", header_cols[2:], row_parts[2:]

    if len(header_upper) > 1 and header_upper[1] == "UNIT":
        unit_name = row_parts[1] if len(row_parts) > 1 else "?"
        if unit_name == "MANL" and len(row_parts) > 2 and row_parts[2] == "PRG":
            return "MANL PRG", header_cols[2:], row_parts[3:]
        if unit_name == "SUB" and len(row_parts) > 2 and row_parts[2] in {"PRO", "PROGRAM"}:
            return "SUB PROGRAM", header_cols[2:], row_parts[3:]
        if unit_name == "LINE" and len(row_parts) > 2:
            suffix = row_parts[2]
            if suffix == "LFT":
                return "LINE LFT", header_cols[2:], row_parts[3:]
            if suffix == "CTR":
                return "LINE CTR", header_cols[2:], row_parts[3:]
            if suffix == "IN":
                return "LINE IN", header_cols[2:], row_parts[3:]
        return unit_name.strip(), header_cols[2:], row_parts[2:]

    return "?", header_cols[1:], row_parts[1:]


def parse_pre_tables(path: Path) -> list[dict]:
    raw = path.read_text(encoding="utf-8", errors="replace")
    if "UNo." not in raw:
        raw = path.read_text(encoding="windows-1252", errors="replace")
    lines = [strip_html(line).rstrip() for line in raw.splitlines() if strip_html(line).strip()]

    blocks: list[dict] = []
    i = 0
    while i < len(lines):
        line = lines[i]
        if not line.strip().startswith("UNo."):
            i += 1
            continue

        header_cols = re.sub(r"\s+", " ", line.strip()).split()
        i += 1
        while i < len(lines) and not lines[i].strip():
            i += 1
        if i >= len(lines):
            break

        row = lines[i]
        i += 1
        row_parts = row.split()
        if not row_parts or not row_parts[0].strip().isdigit():
            continue

        unit_name, data_cols, values = infer_pbd_unit_from_header(header_cols, row_parts)
        if unit_name == "?" or unit_name.isdigit():
            continue

        block = {
            "file": path.name,
            "unit_name": unit_name,
            "header": header_cols,
            "data_cols": data_cols,
            "values": values,
            "sub_units": [],
        }

        while i < len(lines):
            sub = lines[i].strip()
            if sub.startswith("UNo."):
                break
            if sub.startswith("OFS"):
                sub_header = re.sub(r"\s+", " ", sub).split()
                i += 1
                sub_rows: list[list[str]] = []
                while i < len(lines):
                    row_line = lines[i].strip()
                    if row_line.startswith("UNo.") or row_line.startswith("OFS"):
                        break
                    if row_line:
                        sub_rows.append(re.split(r"\s{2,}", row_line))
                    i += 1
                block["sub_units"].append({"kind": "OFS", "header": sub_header, "rows": sub_rows})
                continue
            if sub.startswith("SNo.") or sub.startswith("FIG"):
                sub_header = re.sub(r"\s+", " ", sub).split()
                i += 1
                sub_rows = []
                while i < len(lines):
                    row_line = lines[i].strip()
                    if row_line.startswith("UNo.") or row_line.startswith("SNo.") or row_line.startswith("FIG"):
                        break
                    if row_line:
                        sub_rows.append(re.split(r"\s{2,}", row_line))
                    i += 1
                block["sub_units"].append({"kind": sub_header[0], "header": sub_header, "rows": sub_rows})
                continue
            i += 1

        blocks.append(block)
    return blocks


def collect_blocks() -> list[dict]:
    blocks: list[dict] = []
    for path in sorted(PBD_DIR.glob("*.html")):
        blocks.extend(parse_pre_tables(path))
    return blocks


def scan_binary_types() -> Counter[int]:
    counts: Counter[int] = Counter()
    for path in PBD_DIR.glob("*.PBD"):
        data = path.read_bytes()
        addr = START
        while addr + UNIT_SIZE <= len(data):
            unit_type = data[addr]
            unit_num = data[addr + 2]
            if unit_type == 0 and unit_num == 0:
                break
            counts[unit_type] += 1
            addr += UNIT_SIZE
    return counts


def build_records(blocks: list[dict]) -> tuple[list[dict], list[dict], list[dict]]:
    units: dict[str, dict] = {}
    params: list[dict] = []
    sub_params: list[dict] = []

    for block in blocks:
        unit_name = block["unit_name"]
        uid = UNIT_IDS.get(unit_name, 0)
        if unit_name not in units:
            units[unit_name] = {
                "unit_id": uid,
                "unit_name": unit_name,
                "sample_files": set(),
                "instance_count": 0,
            }
        units[unit_name]["sample_files"].add(block["file"])
        units[unit_name]["instance_count"] += 1

        for col in block["data_cols"]:
            col = col.strip()
            if not col:
                continue
            ptype = PBD_PARAM_TYPES.get(col, infer_param_type(col))
            pos = 0
            for pname, offset, _ in BINARY_OFFSETS.get(uid, []):
                if pname == col or pname.replace(".", "") == col.replace(".", ""):
                    pos = offset
                    break
            params.append(
                {
                    "unit_id": uid,
                    "unit_name": unit_name,
                    "parameter": col,
                    "display_type": ptype,
                    "binary_offset": pos,
                    "source": "binary-validated" if pos else "html-inferred",
                    "sample_file": block["file"],
                }
            )

        for sub in block["sub_units"]:
            kind = sub["kind"]
            if kind == "OFS":
                continue
            parent = unit_name
            if kind == "FIG":
                sid = {
                    "LINE LFT": SUB_UNIT_IDS["FIG-LINE"],
                    "LINE CTR": SUB_UNIT_IDS["FIG-LINE"],
                    "LINE IN": SUB_UNIT_IDS["FIG-LINE"],
                    "POCKET": SUB_UNIT_IDS["FIG-POCKET"],
                    "DRILLING": SUB_UNIT_IDS["FIG-DRILLING"],
                    "CIRC MIL": SUB_UNIT_IDS["FIG-DRILLING"],
                    "REAMING": SUB_UNIT_IDS["FIG-DRILLING"],
                }.get(parent, SUB_UNIT_IDS["FIG-LINE"])
                sub_name = f"FIG ({parent})"
            else:
                sid = {
                    "LINE LFT": SUB_UNIT_IDS["SNo-LINE"],
                    "LINE CTR": SUB_UNIT_IDS["SNo-LINE"],
                    "LINE IN": SUB_UNIT_IDS["SNo-LINE"],
                    "POCKET": SUB_UNIT_IDS["SNo-POCKET"],
                    "DRILLING": SUB_UNIT_IDS["SNo-DRILLING"],
                    "CIRC MIL": SUB_UNIT_IDS["SNo-DRILLING"],
                    "REAMING": SUB_UNIT_IDS["SNo-DRILLING"],
                    "MANL PRG": SUB_UNIT_IDS["SNo-MANL"],
                }.get(parent, SUB_UNIT_IDS["SNo-LINE"])
                sub_name = f"SNo ({parent})"

            header = [h for h in sub["header"] if h and h not in {"FIG", "SNo."}]
            for col in header:
                ptype = PBD_PARAM_TYPES.get(col, infer_param_type(col))
                pos = 0
                for pname, offset, _ in BINARY_OFFSETS.get(sid, []):
                    if pname == col:
                        pos = offset
                        break
                sub_params.append(
                    {
                        "sub_unit_id": sid,
                        "sub_unit_name": sub_name,
                        "parent_unit": parent,
                        "parameter": col,
                        "display_type": ptype,
                        "binary_offset": pos,
                        "source": "binary-validated" if pos else "html-inferred",
                        "sample_file": block["file"],
                    }
                )

    unit_rows = []
    for name in sorted(units, key=lambda n: units[n]["unit_id"] or 999):
        u = units[name]
        unit_rows.append(
            {
                "unit_id": u["unit_id"],
                "unit_name": name,
                "instance_count": u["instance_count"],
                "sample_files": ", ".join(sorted(u["sample_files"])),
                "notes": "" if u["unit_id"] else "Unknown ID — check binary scan",
            }
        )

    return (
        unit_rows,
        dedupe_rows(params, ["unit_name", "parameter"]),
        dedupe_rows(sub_params, ["sub_unit_name", "parameter"]),
    )


def write_xlsx(unit_rows: list[dict], param_rows: list[dict], sub_rows: list[dict]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Units"
    ws.append(["unit_id", "unit_name", "instance_count", "sample_files", "notes"])
    for row in unit_rows:
        ws.append([row["unit_id"], row["unit_name"], row["instance_count"], row["sample_files"], row["notes"]])

    ws2 = wb.create_sheet("Parameters")
    ws2.append(["unit_id", "unit_name", "parameter", "display_type", "binary_offset", "source"])
    for row in param_rows:
        ws2.append(
            [row["unit_id"], row["unit_name"], row["parameter"], row["display_type"], row["binary_offset"], row["source"]]
        )

    ws3 = wb.create_sheet("SubUnits")
    ws3.append(["sub_unit_id", "sub_unit_name", "parent_unit", "parameter", "display_type", "binary_offset", "source"])
    for row in sub_rows:
        ws3.append(
            [
                row["sub_unit_id"],
                row["sub_unit_name"],
                row["parent_unit"],
                row["parameter"],
                row["display_type"],
                row["binary_offset"],
                row["source"],
            ]
        )

    ws4 = wb.create_sheet("BinaryScan")
    ws4["A1"] = "Unit type frequency in SAMPLE_NC_PROGRAM/PBD/*.PBD"
    ws4["A1"].font = Font(bold=True)
    ws4.append(["unit_id", "count"])
    for unit_id, count in sorted(scan_binary_types().items()):
        ws4.append([unit_id, count])

    ws5 = wb.create_sheet("About")
    ws5["A1"] = "PBD structure reverse-engineered from SAMPLE_NC_PROGRAM/PBD HTML + binary validation"
    ws5["A1"].font = Font(bold=True)
    ws5["A3"] = "Matrix milling programs: LINE LFT/CTR/IN, DRILLING, CIRC MIL, REAMING, POCKET, INDEX, WPC-."
    ws5["A4"] = "Hidden binary slots: type 160 (OFS offset table), type 222 (internal metadata)."

    wb.save(OUT_XLSX)


def append_unit(root: ET.Element, unit_id: int, name: str, params: list[tuple[str, int, str]]) -> None:
    unit = ET.SubElement(root, "unit", id=str(unit_id), name=name)
    for pname, pos, ptype in params:
        param = ET.SubElement(unit, "parameter", name=pname, pos=str(pos), type=ptype)
        if pname == "PTN" and unit_id in PBD_PTN_ENUMS:
            for enum_name, enum_value in PBD_PTN_ENUMS[unit_id]:
                ET.SubElement(param, "enum", name=enum_name, value=str(enum_value), type="")
        if pname == "TOOL" and ptype == "readPbdTool":
            for enum_name, enum_value in PBD_TOOL_ENUMS:
                ET.SubElement(param, "enum", name=enum_name, value=str(enum_value), type="")


def ensure_ptn_enums(root: ET.Element) -> None:
    for unit in root.findall("unit"):
        unit_id = unit.get("id")
        if unit_id is None:
            continue
        uid = int(unit_id)
        if uid not in PBD_PTN_ENUMS:
            continue
        for param in unit.findall("parameter"):
            if param.get("name") != "PTN":
                continue
            existing = {e.get("value") for e in param.findall("enum")}
            for enum_name, enum_value in PBD_PTN_ENUMS[uid]:
                if str(enum_value) not in existing:
                    ET.SubElement(param, "enum", name=enum_name, value=str(enum_value), type="")


def ensure_pbd_tool_enums(root: ET.Element) -> None:
    for unit in root.findall("unit"):
        if unit.get("name") != "SNo":
            continue
        for param in unit.findall("parameter"):
            if param.get("name") != "TOOL":
                continue
            param.set("pos", "9")
            param.set("type", "readPbdTool")
            existing = {e.get("value") for e in param.findall("enum")}
            for enum_name, enum_value in PBD_TOOL_ENUMS:
                if str(enum_value) not in existing:
                    ET.SubElement(param, "enum", name=enum_name, value=str(enum_value), type="")


def ensure_sno_m_columns(root: ET.Element) -> None:
    """PBD SNo rows expose three M-code columns at binary offsets 24, 26, and 32."""
    for unit_id in (176, 177, 178):
        for unit in root.findall("unit"):
            if unit.get("id") != str(unit_id):
                continue
            existing = [p for p in unit.findall("parameter") if p.get("name") == "M"]
            if len(existing) == 3 and [int(p.get("pos", "0")) for p in existing] == [24, 26, 32]:
                continue
            for param in existing:
                unit.remove(param)
            insert_at = len(list(unit))
            for child in unit.findall("parameter"):
                if child.get("name") == "FR":
                    insert_at = list(unit).index(child) + 1
            for offset in (24, 26, 32):
                param = ET.Element(
                    "parameter",
                    {"name": "M", "pos": str(offset), "type": "readFullNumber2B"},
                )
                unit.insert(insert_at, param)
                insert_at += 1


def ensure_pbd_mat_unit(root: ET.Element) -> None:
    """Replace turning MAT layout with Matrix PBD header columns."""
    for unit in root.findall("unit"):
        if unit.get("id") != "1":
            continue
        for param in list(unit.findall("parameter")):
            unit.remove(param)
        for pname, pos, ptype in BINARY_OFFSETS[1]:
            param = ET.SubElement(unit, "parameter", name=pname, pos=str(pos), type=ptype)
            if pname == "MULTI MODE":
                for enum_name, enum_value in PBD_MULTI_MODE_ENUMS:
                    ET.SubElement(param, "enum", name=enum_name, value=str(enum_value), type="")


def ensure_ofs_unit(root: ET.Element) -> None:
    """Ensure type-160 offset table (OFS row under MAT) exists."""
    for unit in root.findall("unit"):
        if unit.get("id") == "160":
            unit.set("name", "OFS")
            existing = {p.get("name") for p in unit.findall("parameter")}
            for pname, pos, ptype in BINARY_OFFSETS[160]:
                if pname not in existing:
                    ET.SubElement(unit, "parameter", name=pname, pos=str(pos), type=ptype)
            return
    append_unit(root, 160, "OFS", BINARY_OFFSETS[160])


def write_xml() -> None:
    shutil.copy(QTS_XML, OUT_XML)
    root = ET.parse(OUT_XML).getroot()
    assert root is not None

    existing_ids = {unit.get("id") for unit in root.findall("unit")}

    name_map = {
        32: "DRILLING",
        35: "REAMING",
        38: "CIRC MIL",
        64: "LINE CTR",
        66: "LINE LFT",
        68: "LINE IN",
        99: "POCKET",
        176: "SNo",
        177: "SNo",
        178: "SNo",
        192: "FIG",
        193: "FIG",
        194: "FIG",
        160: "OFS",
    }

    for unit_id, params in sorted(BINARY_OFFSETS.items()):
        if str(unit_id) in existing_ids:
            for unit in root.findall("unit"):
                if unit.get("id") != str(unit_id):
                    continue
                if unit_id in name_map:
                    unit.set("name", name_map[unit_id])
                existing_names = {p.get("name") for p in unit.findall("parameter")}
                for pname, pos, ptype in params:
                    if pname not in existing_names:
                        ET.SubElement(unit, "parameter", name=pname, pos=str(pos), type=ptype)
            continue

        append_unit(root, unit_id, name_map.get(unit_id, "TBD"), params)

    ensure_pbd_mat_unit(root)
    ensure_ofs_unit(root)

    # Extend WPC- coordinates.
    for unit in root.findall("unit"):
        if unit.get("id") == "2":
            existing = {p.get("name") for p in unit.findall("parameter")}
            for pname, pos, ptype in BINARY_OFFSETS[2]:
                if pname not in existing:
                    ET.SubElement(unit, "parameter", name=pname, pos=str(pos), type=ptype)

    # Extend INDEX.
    for unit in root.findall("unit"):
        if unit.get("id") == "12":
            existing = {p.get("name") for p in unit.findall("parameter")}
            for pname, pos, ptype in BINARY_OFFSETS[12]:
                if pname not in existing:
                    ET.SubElement(unit, "parameter", name=pname, pos=str(pos), type=ptype)

    ensure_ptn_enums(root)
    ensure_pbd_tool_enums(root)
    ensure_sno_m_columns(root)

    tree = ET.ElementTree(root)
    ET.indent(tree, space="    ")
    tree.write(OUT_XML, encoding="utf-8", xml_declaration=True)


def main() -> None:
    blocks = collect_blocks()
    unit_rows, param_rows, sub_rows = build_records(blocks)
    write_xlsx(unit_rows, param_rows, sub_rows)
    write_xml()
    print(f"Parsed {len(blocks)} unit blocks from {len(list(PBD_DIR.glob('*.html')))} HTML files")
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_XML}")
    print(f"Distinct unit types: {len(unit_rows)}")


if __name__ == "__main__":
    main()
