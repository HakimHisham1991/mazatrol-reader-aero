"""Build m6m_structure.xlsx and m6m_structure.xml from SAMPLE_NC_PROGRAM/M6M HTML + binary."""

from __future__ import annotations

import re
import shutil
import struct
import sys
import xml.etree.ElementTree as ET
from collections import Counter
from html.parser import HTMLParser
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from tools.build_pbd_structure import (  # noqa: E402
    PBD_MULTI_MODE_ENUMS,
    PBD_TOOL_ENUMS,
    dedupe_rows,
    ensure_pbd_tool_enums,
    ensure_sno_m_columns,
)
from tools.build_pbf_structure import PARAM_TYPES, infer_param_type, strip_html  # noqa: E402

M6M_DIR = ROOT / "SAMPLE_NC_PROGRAM" / "M6M"
PBF_XML = ROOT / "pbf_structure.xml"
OUT_XLSX = ROOT / "m6m_structure.xlsx"
OUT_XML = ROOT / "m6m_structure.xml"

START = 0xFC
UNIT_SIZE = 100
M6M_MAT_NUM = 251

UNIT_IDS: dict[str, int] = {
    "MAT": 1,
    "WPC-": 2,
    "END": 4,
    "SUB PRO": 5,
    "SUB PROGRAM": 5,
    "MANU PRO": 6,
    "MANL PRG": 6,
    "DRILLING": 32,
    "CIRC MIL": 38,
    "FACE MIL": 96,
    "POCKET": 99,
    "LINE LFT": 66,
    "LINE OUT": 67,
    "TAPPING": 55,
}

SUB_UNIT_IDS: dict[str, int] = {
    "SNo-MANL": 161,
    "SNo-MILL": 177,
    "SNo-POCKET": 178,
    "SNo-DRILL": 176,
    "FIG-CONTOUR": 194,
    "FIG-RECT": 193,
    "FIG-DRILL": 192,
}

M6M_PTN_ENUMS: dict[int, list[tuple[str, int]]] = {
    192: [("PT", 1)],
    193: [("SQR", 16), ("LINE", 32), ("ARC", 33), ("CW", 33), ("CCW", 34)],
    194: [("LINE", 32), ("CW", 33), ("CCW", 34)],
}

BINARY_OFFSETS: dict[int, list[tuple[str, int, str]]] = {
    1: [
        ("MAT.", 4, "text"),
        ("INITIAL-Z", 16, "readData"),
        ("ATC MODE", 10, "readFullNumber2B"),
        ("MULTI MODE", 9, "readPattern"),
        ("MULTI FLAG", 9, "readPbdMultiFlag"),
        ("PITCH-X", 72, "readData"),
        ("PITCH-Y", 76, "readData"),
    ],
    2: [
        ("ADD. WPC", 32, "readFullNumber2B"),
        ("X", 66, "readData"),
        ("Y", 70, "readData"),
        ("th", 74, "readData"),
        ("Z", 78, "readData"),
    ],
    4: [
        ("CONTI.", 10, "readFullNumber2B"),
        ("REPEAT", 0, "readFullNumber2B"),
        ("NUMBER", 0, "readFullNumber2B"),
        ("RETURN", 0, "readFullNumber2B"),
    ],
    5: [
        ("WRK No.", 36, "readData"),
        ("REPEAT", 40, "readFullNumber2B"),
    ],
    6: [
        ("TOOL", 84, "text"),
        ("NOM-Ø", 64, "readData"),
        ("No.", 68, "readFullNumber2B"),
    ],
    32: [
        ("DIA", 36, "readData"),
        ("DEPTH", 40, "readData"),
        ("CHMF", 44, "readData"),
    ],
    38: [
        ("TORNAD", 32, "readFullNumber2B"),
        ("DIA", 36, "readData"),
        ("DEPTH", 40, "readData"),
        ("CHMF", 44, "readData"),
        ("BTM", 48, "readFullNumber2B"),
        ("PITCH1", 52, "readData"),
        ("PITCH2", 56, "readData"),
    ],
    55: [
        ("NOM-", 36, "readData"),
        ("MAJOR-Ø", 40, "readData"),
        ("PITCH", 44, "readData"),
        ("TAP-DEP", 48, "readData"),
        ("CHMF", 52, "readData"),
        ("CHP", 56, "readFullNumber2B"),
    ],
    66: [
        ("DEPTH", 36, "readData"),
        ("SRV-Z", 40, "readData"),
        ("SRV-R", 44, "readData"),
        ("RGH", 48, "readFullNumber2B"),
        ("FIN-Z", 52, "readData"),
        ("FIN-R", 56, "readData"),
    ],
    67: [
        ("DEPTH", 36, "readData"),
        ("SRV-Z", 40, "readData"),
        ("SRV-R", 44, "readData"),
        ("RGH", 48, "readFullNumber2B"),
        ("FIN-Z", 52, "readData"),
        ("FIN-R", 56, "readData"),
    ],
    96: [
        ("DEPTH", 36, "readData"),
        ("SRV-Z", 40, "readData"),
        ("BTM", 48, "readFullNumber2B"),
        ("WAL", 52, "readFullNumber2B"),
        ("FIN-Z", 56, "readData"),
        ("FIN-R", 60, "readData"),
    ],
    99: [
        ("DEPTH", 36, "readData"),
        ("SRV-Z", 40, "readData"),
        ("BTM", 48, "readFullNumber2B"),
        ("WAL", 52, "readFullNumber2B"),
        ("FIN-Z", 56, "readData"),
        ("FIN-R", 60, "readData"),
        ("INTER-R", 64, "readData"),
        ("CHMF", 68, "readData"),
    ],
    161: [
        ("G0", 20, "readFullNumber2B"),
        ("G1", 22, "readFullNumber2B"),
        ("G2", 24, "readFullNumber2B"),
        ("DATA1_L", 8, "readLetter"),
        ("DATA1_V", 36, "readData"),
        ("DATA2_L", 9, "readLetter"),
        ("DATA2_V", 40, "readData"),
        ("DATA3_L", 10, "readLetter"),
        ("DATA3_V", 44, "readData"),
        ("DATA4_L", 11, "readLetter"),
        ("DATA4_V", 48, "readData"),
        ("DATA5_L", 12, "readLetter"),
        ("DATA5_V", 52, "readData"),
        ("DATA6_L", 13, "readLetter"),
        ("DATA6_V", 56, "readData"),
        ("M/B", 24, "readFullNumber2B"),
    ],
    176: [
        ("TOOL", 9, "readPbdTool"),
        ("NOM-Ø", 36, "readData"),
        ("HOLE-Ø", 40, "readData"),
        ("HOLE-DEP", 44, "readData"),
        ("PRE-DIA", 48, "readData"),
        ("PRE-DEP", 52, "readData"),
        ("RGH", 56, "readFullNumber2B"),
        ("DEPTH", 60, "readData"),
        ("C-SP", 64, "readFullNumber2B"),
        ("FR", 68, "readData"),
        ("M", 24, "readFullNumber2B"),
        ("M", 26, "readFullNumber2B"),
    ],
    177: [
        ("TOOL", 9, "readPbdTool"),
        ("NOM-Ø", 36, "readData"),
        ("No.", 68, "readFullNumber2B"),
        ("APRCH-X", 40, "readData"),
        ("APRCH-Y", 44, "readData"),
        ("TYPE", 48, "readPattern"),
        ("ZFD", 52, "readFullNumber2B"),
        ("DEP-Z", 56, "readData"),
        ("WID-R", 60, "readData"),
        ("C-SP", 64, "readFullNumber2B"),
        ("FR", 68, "readData"),
        ("M", 24, "readFullNumber2B"),
        ("M", 26, "readFullNumber2B"),
    ],
    178: [
        ("TOOL", 9, "readPbdTool"),
        ("NOM-Ø", 36, "readData"),
        ("PK-DEP", 48, "readData"),
        ("WID-R", 52, "readData"),
        ("C-SP", 60, "readFullNumber2B"),
        ("FR", 64, "readData"),
        ("M", 24, "readFullNumber2B"),
        ("M", 26, "readFullNumber2B"),
    ],
    192: [
        ("PTN", 8, "readPattern"),
        ("Z", 36, "readData"),
        ("X", 40, "readData"),
        ("Y", 44, "readData"),
        ("AN1", 48, "readData"),
        ("AN2", 52, "readData"),
        ("T1", 56, "readData"),
        ("T2", 60, "readData"),
    ],
    193: [
        ("PTN", 8, "readPattern"),
        ("P1X/CX", 36, "readData"),
        ("P1Y/CY", 40, "readData"),
        ("P3X/R", 44, "readData"),
        ("P3Y", 48, "readData"),
        ("CN1", 52, "readData"),
        ("CN2", 56, "readData"),
        ("CN3", 60, "readData"),
        ("CN4", 64, "readData"),
    ],
    194: [
        ("PTN", 8, "readPattern"),
        ("X", 40, "readData"),
        ("Y", 44, "readData"),
        ("R/th", 56, "readData"),
        ("I", 60, "readData"),
        ("J", 64, "readData"),
        ("CNR", 68, "readData"),
    ],
}

M6M_PARAM_TYPES = {
    **PARAM_TYPES,
    "MAT.": "text",
    "INITIAL-Z": "readData",
    "MULTI MODE": "readPattern",
    "MULTI FLAG": "readPbdMultiFlag",
    "PITCH-X": "readData",
    "PITCH-Y": "readData",
    "MAJOR-Ø": "readData",
    "TAP-DEP": "readData",
    "CHP": "readFullNumber2B",
    "NOM-": "readData",
    "P1X/CX": "readData",
    "P1Y/CY": "readData",
    "P3X/R": "readData",
    "P3Y": "readData",
    "CN1": "readData",
    "CN2": "readData",
    "CN3": "readData",
    "CN4": "readData",
    "M/B": "readFullNumber2B",
    "DATA 1": "readLetter",
    "DATA 2": "readLetter",
    "DATA 3": "readLetter",
    "DATA 4": "readLetter",
    "DATA 5": "readLetter",
    "DATA 6": "readLetter",
}


def infer_m6m_unit_from_header(header_cols: list[str], row_parts: list[str]) -> tuple[str, list[str], list[str]]:
    header_upper = [c.upper().replace(".", "") for c in header_cols]
    if header_upper[0] != "UNO":
        return "?", header_cols[1:], row_parts

    if len(header_upper) > 1 and header_upper[1] in {"MAT", "MATERIAL"}:
        return "MAT", header_cols[2:], row_parts[1:]

    if "INITIAL-Z" in header_upper:
        return "MAT", header_cols[2:], row_parts[1:]

    if "ADD. WPC" in header_upper or "ADD WPC" in " ".join(header_upper):
        return "WPC-", header_cols[2:], row_parts[2:]

    if "WORK NO" in " ".join(header_upper) or ("REPEAT" in header_upper and "CONTI" not in header_upper):
        return "SUB PRO", header_cols[2:], row_parts[2:]

    if "CONTI" in header_upper and "ATC" in header_upper:
        return "END", header_cols[2:], row_parts[2:]

    if len(header_upper) > 1 and header_upper[1] == "UNIT":
        unit_name = row_parts[1] if len(row_parts) > 1 else "?"
        if unit_name in {"MANU", "MANL"} and len(row_parts) > 2 and row_parts[2] in {"PRO", "PRG"}:
            return "MANU PRO", header_cols[2:], row_parts[3:]
        if unit_name == "SUB" and len(row_parts) > 2 and row_parts[2] in {"PRO", "PROGRAM"}:
            return "SUB PRO", header_cols[2:], row_parts[3:]
        if unit_name == "LINE" and len(row_parts) > 2:
            suffix = row_parts[2]
            if suffix == "LFT":
                return "LINE LFT", header_cols[2:], row_parts[3:]
            if suffix == "OUT":
                return "LINE OUT", header_cols[2:], row_parts[3:]
        if unit_name.startswith("TAPPING"):
            return "TAPPING", header_cols[2:], row_parts[2:]
        return unit_name.strip(), header_cols[2:], row_parts[2:]

    if "SRV-Z" in header_upper and "BTM" in header_upper and "WAL" in header_upper:
        unit_name = row_parts[1] if len(row_parts) > 1 else "POCKET"
        if unit_name.strip() == "FACE MIL":
            return "FACE MIL", header_cols[2:], row_parts[2:]
        return "POCKET", header_cols[2:], row_parts[2:]

    if "SRV-Z" in header_upper and "SRV-R" in header_upper:
        unit_name = row_parts[1] if len(row_parts) > 1 else "LINE LFT"
        if len(row_parts) > 2 and row_parts[2] == "OUT":
            return "LINE OUT", header_cols[2:], row_parts[3:]
        if len(row_parts) > 2 and row_parts[2] == "LFT":
            return "LINE LFT", header_cols[2:], row_parts[3:]
        return unit_name.strip(), header_cols[2:], row_parts[2:]

    if "DIA" in header_upper and "DEPTH" in header_upper:
        if "TORNAD" in header_upper:
            return "CIRC MIL", header_cols[2:], row_parts[2:]
        if "MAJOR" in " ".join(header_upper) or "TAP-DEP" in " ".join(header_upper):
            return "TAPPING", header_cols[2:], row_parts[2:]
        return "DRILLING", header_cols[2:], row_parts[2:]

    if "TOOL" in header_upper and "NOM" in " ".join(header_upper) and "UNIT" in header_upper:
        return "MANU PRO", header_cols[2:], row_parts[2:]

    return "?", header_cols[1:], row_parts[1:]


def parse_pre_tables(path: Path) -> list[dict]:
    raw = path.read_text(encoding="utf-8", errors="replace")
    if "UNo." not in raw:
        raw = path.read_text(encoding="windows-1252", errors="replace")
    lines = [strip_html(line).rstrip() for line in raw.splitlines() if strip_html(line).strip()]

    blocks: list[dict] = []
    i = 0
    while i < len(lines):
        line = lines[i]
        if not line.strip().startswith("UNo."):
            i += 1
            continue

        header_cols = re.sub(r"\s+", " ", line.strip()).split()
        i += 1
        while i < len(lines) and not lines[i].strip():
            i += 1
        if i >= len(lines):
            break

        row = lines[i]
        i += 1
        row_parts = row.split()
        if not row_parts or not row_parts[0].strip().isdigit():
            continue

        unit_name, data_cols, values = infer_m6m_unit_from_header(header_cols, row_parts)
        if unit_name == "?" or unit_name.isdigit():
            continue

        block = {
            "file": path.name,
            "unit_name": unit_name,
            "header": header_cols,
            "data_cols": data_cols,
            "values": values,
            "sub_units": [],
        }

        while i < len(lines):
            sub = lines[i].strip()
            if sub.startswith("UNo."):
                break
            if sub.startswith("SNo") or sub.startswith("FIG"):
                sub_header = re.sub(r"\s+", " ", sub).split()
                i += 1
                sub_rows: list[list[str]] = []
                while i < len(lines):
                    row_line = lines[i].strip()
                    if row_line.startswith("UNo.") or row_line.startswith("SNo") or row_line.startswith("FIG"):
                        break
                    if row_line:
                        sub_rows.append(re.split(r"\s{2,}", row_line))
                    i += 1
                block["sub_units"].append({"kind": sub_header[0], "header": sub_header, "rows": sub_rows})
                continue
            i += 1

        blocks.append(block)
    return blocks


def collect_blocks() -> list[dict]:
    blocks: list[dict] = []
    for path in sorted(M6M_DIR.glob("*.html")):
        blocks.extend(parse_pre_tables(path))
    return blocks


def scan_binary_types() -> Counter[int]:
    counts: Counter[int] = Counter()
    for path in M6M_DIR.glob("*.M6M"):
        data = path.read_bytes()
        addr = START
        while addr + UNIT_SIZE <= len(data):
            unit_type = data[addr]
            unit_num = data[addr + 2]
            if unit_type == 0 and unit_num == 0:
                break
            counts[unit_type] += 1
            addr += UNIT_SIZE
    return counts


def build_records(blocks: list[dict]) -> tuple[list[dict], list[dict], list[dict]]:
    units: dict[str, dict] = {}
    params: list[dict] = []
    sub_params: list[dict] = []

    for block in blocks:
        unit_name = block["unit_name"]
        uid = UNIT_IDS.get(unit_name, 0)
        if unit_name not in units:
            units[unit_name] = {
                "unit_id": uid,
                "unit_name": unit_name,
                "sample_files": set(),
                "instance_count": 0,
            }
        units[unit_name]["sample_files"].add(block["file"])
        units[unit_name]["instance_count"] += 1

        for col in block["data_cols"]:
            col = col.strip()
            if not col:
                continue
            ptype = M6M_PARAM_TYPES.get(col, infer_param_type(col))
            pos = 0
            for pname, offset, _ in BINARY_OFFSETS.get(uid, []):
                if pname == col or pname.replace(".", "") == col.replace(".", ""):
                    pos = offset
                    break
            params.append(
                {
                    "unit_id": uid,
                    "unit_name": unit_name,
                    "parameter": col,
                    "display_type": ptype,
                    "binary_offset": pos,
                    "source": "binary-validated" if pos else "html-inferred",
                    "sample_file": block["file"],
                }
            )

        for sub in block["sub_units"]:
            kind = sub["kind"]
            parent = unit_name
            header_join = " ".join(sub["header"])
            if kind.startswith("FIG"):
                if "P1X" in header_join or "P1Y" in header_join:
                    sid = SUB_UNIT_IDS["FIG-RECT"]
                    sub_name = f"FIG-RECT ({parent})"
                elif "Z" in header_join and "AN1" in header_join:
                    sid = SUB_UNIT_IDS["FIG-DRILL"]
                    sub_name = f"FIG-DRILL ({parent})"
                else:
                    sid = SUB_UNIT_IDS["FIG-CONTOUR"]
                    sub_name = f"FIG-CONTOUR ({parent})"
            elif kind.startswith("SNo"):
                if parent in {"MANU PRO", "MANL PRG"}:
                    sid = SUB_UNIT_IDS["SNo-MANL"]
                    sub_name = f"SNo-MANL ({parent})"
                elif parent == "POCKET":
                    sid = SUB_UNIT_IDS["SNo-POCKET"]
                    sub_name = f"SNo-POCKET ({parent})"
                elif parent in {"DRILLING", "CIRC MIL", "TAPPING"}:
                    sid = SUB_UNIT_IDS["SNo-DRILL"]
                    sub_name = f"SNo-DRILL ({parent})"
                else:
                    sid = SUB_UNIT_IDS["SNo-MILL"]
                    sub_name = f"SNo-MILL ({parent})"
            else:
                continue

            header = [h for h in sub["header"] if h and h not in {"FIG", "SNo.", "SNo.G1", "SNo"}]
            for col in header:
                ptype = M6M_PARAM_TYPES.get(col, infer_param_type(col))
                pos = 0
                for pname, offset, _ in BINARY_OFFSETS.get(sid, []):
                    if pname == col or pname.replace(".", "") == col.replace(".", ""):
                        pos = offset
                        break
                sub_params.append(
                    {
                        "sub_unit_id": sid,
                        "sub_unit_name": sub_name,
                        "parent_unit": parent,
                        "parameter": col,
                        "display_type": ptype,
                        "binary_offset": pos,
                        "source": "binary-validated" if pos else "html-inferred",
                        "sample_file": block["file"],
                    }
                )

    unit_rows = []
    for name in sorted(units, key=lambda n: units[n]["unit_id"] or 999):
        u = units[name]
        unit_rows.append(
            {
                "unit_id": u["unit_id"],
                "unit_name": name,
                "instance_count": u["instance_count"],
                "sample_files": ", ".join(sorted(u["sample_files"])),
                "notes": "" if u["unit_id"] else "Unknown ID — check binary scan",
            }
        )

    return (
        unit_rows,
        dedupe_rows(params, ["unit_name", "parameter"]),
        dedupe_rows(sub_params, ["sub_unit_name", "parameter"]),
    )


def write_xlsx(unit_rows: list[dict], param_rows: list[dict], sub_rows: list[dict]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Units"
    ws.append(["unit_id", "unit_name", "instance_count", "sample_files", "notes"])
    for row in unit_rows:
        ws.append([row["unit_id"], row["unit_name"], row["instance_count"], row["sample_files"], row["notes"]])

    ws2 = wb.create_sheet("Parameters")
    ws2.append(["unit_id", "unit_name", "parameter", "display_type", "binary_offset", "source"])
    for row in param_rows:
        ws2.append(
            [row["unit_id"], row["unit_name"], row["parameter"], row["display_type"], row["binary_offset"], row["source"]]
        )

    ws3 = wb.create_sheet("SubUnits")
    ws3.append(["sub_unit_id", "sub_unit_name", "parent_unit", "parameter", "display_type", "binary_offset", "source"])
    for row in sub_rows:
        ws3.append(
            [
                row["sub_unit_id"],
                row["sub_unit_name"],
                row["parent_unit"],
                row["parameter"],
                row["display_type"],
                row["binary_offset"],
                row["source"],
            ]
        )

    ws4 = wb.create_sheet("BinaryScan")
    ws4["A1"] = "Unit type frequency in SAMPLE_NC_PROGRAM/M6M/*.M6M @ 0xFC stride"
    ws4["A1"].font = Font(bold=True)
    ws4.append(["unit_id", "count"])
    for unit_id, count in sorted(scan_binary_types().items()):
        ws4.append([unit_id, count])

    ws5 = wb.create_sheet("About")
    ws5["A1"] = "M6M / M640M milling programs — reverse-engineered from HTML + binary"
    ws5["A1"].font = Font(bold=True)
    ws5["A3"] = "MAT slot: type=0 num=251 @ 0xFC. WPC coords embedded in MAT block (+66..+78)."
    ws5["A4"] = "Mill units: POCKET=99, LINE LFT=66, LINE OUT=67, FACE MIL=96. Type-0 num=6/62 = SNo/FIG."

    wb.save(OUT_XLSX)


def append_unit(root: ET.Element, unit_id: int, name: str, params: list[tuple[str, int, str]]) -> None:
    unit = ET.SubElement(root, "unit", id=str(unit_id), name=name)
    for pname, pos, ptype in params:
        param = ET.SubElement(unit, "parameter", name=pname, pos=str(pos), type=ptype)
        if pname == "PTN" and unit_id in M6M_PTN_ENUMS:
            for enum_name, enum_value in M6M_PTN_ENUMS[unit_id]:
                ET.SubElement(param, "enum", name=enum_name, value=str(enum_value), type="")
        if pname == "MULTI MODE":
            for enum_name, enum_value in PBD_MULTI_MODE_ENUMS:
                ET.SubElement(param, "enum", name=enum_name, value=str(enum_value), type="")
        if pname == "TOOL" and ptype == "readPbdTool":
            for enum_name, enum_value in PBD_TOOL_ENUMS:
                ET.SubElement(param, "enum", name=enum_name, value=str(enum_value), type="")


def ensure_ptn_enums(root: ET.Element) -> None:
    for unit in root.findall("unit"):
        unit_id = unit.get("id")
        if unit_id is None:
            continue
        uid = int(unit_id)
        if uid not in M6M_PTN_ENUMS:
            continue
        for param in unit.findall("parameter"):
            if param.get("name") != "PTN":
                continue
            existing = {e.get("value") for e in param.findall("enum")}
            for enum_name, enum_value in M6M_PTN_ENUMS[uid]:
                if str(enum_value) not in existing:
                    ET.SubElement(param, "enum", name=enum_name, value=str(enum_value), type="")


def ensure_m6m_units(root: ET.Element) -> None:
    name_map = {
        1: "MAT",
        2: "WPC-",
        4: "END",
        5: "SUB PRO",
        6: "MANU PRO",
        32: "DRILLING",
        38: "CIRC MIL",
        55: "TAPPING",
        66: "LINE LFT",
        67: "LINE OUT",
        96: "FACE MIL",
        99: "POCKET",
        161: "SNo",
        176: "SNo",
        177: "SNo",
        178: "SNo",
        192: "FIG",
        193: "FIG",
        194: "FIG",
    }
    existing_ids = {unit.get("id") for unit in root.findall("unit")}

    for unit_id, params in sorted(BINARY_OFFSETS.items()):
        if str(unit_id) in existing_ids:
            for unit in root.findall("unit"):
                if unit.get("id") != str(unit_id):
                    continue
                unit.set("name", name_map.get(unit_id, unit.get("name", "TBD")))
                for param in list(unit.findall("parameter")):
                    unit.remove(param)
                for pname, pos, ptype in params:
                    param = ET.SubElement(unit, "parameter", name=pname, pos=str(pos), type=ptype)
                    if pname == "MULTI MODE":
                        for enum_name, enum_value in PBD_MULTI_MODE_ENUMS:
                            ET.SubElement(param, "enum", name=enum_name, value=str(enum_value), type="")
                    if pname == "TOOL" and ptype == "readPbdTool":
                        for enum_name, enum_value in PBD_TOOL_ENUMS:
                            ET.SubElement(param, "enum", name=enum_name, value=str(enum_value), type="")
                    if pname == "PTN" and unit_id in M6M_PTN_ENUMS:
                        for enum_name, enum_value in M6M_PTN_ENUMS[unit_id]:
                            ET.SubElement(param, "enum", name=enum_name, value=str(enum_value), type="")
            continue
        append_unit(root, unit_id, name_map.get(unit_id, "TBD"), params)


def write_xml() -> None:
    shutil.copy(PBF_XML, OUT_XML)
    root = ET.parse(OUT_XML).getroot()
    assert root is not None
    ensure_m6m_units(root)
    ensure_ptn_enums(root)
    ensure_pbd_tool_enums(root)
    ensure_sno_m_columns(root)
    tree = ET.ElementTree(root)
    ET.indent(tree, space="    ")
    tree.write(OUT_XML, encoding="utf-8", xml_declaration=True)


def main() -> None:
    blocks = collect_blocks()
    unit_rows, param_rows, sub_rows = build_records(blocks)
    write_xlsx(unit_rows, param_rows, sub_rows)
    write_xml()
    print(f"Parsed {len(blocks)} unit blocks from {len(list(M6M_DIR.glob('*.html')))} HTML files")
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_XML}")
    print(f"Distinct unit types: {len(unit_rows)}")


if __name__ == "__main__":
    main()
